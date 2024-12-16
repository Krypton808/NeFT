import evaluate
import csv
import jsonlines
import openpyxl
from openpyxl import Workbook

from rouge_score import rouge_scorer
from rouge.rouge import FilesRouge


def Rouge(predictions_path, references_path):
    predictions_list = []
    references_list = []
    pf = open(predictions_path, 'r', encoding='utf-8')
    r_lines = jsonlines.open(references_path)

    p_rows = csv.reader(pf)

    for row in p_rows:
        predictions_list.append(row[0])

    for line in r_lines:
        references_list.append(line['summary'].strip())

    # print(predictions_list)
    # print(references_list)

    # predictions_list = ["hello there", "general kenobi"]
    # references_list = ["hello there", "general kenobi"]

    rouge = evaluate.load(r'/mnt/nfs/algo/intern/haoyunx9/evaluate_metric/rouge')
    results = rouge.compute(predictions=predictions_list, references=references_list)

    print(results)
    print('+++++++++++++++++++++++++++++++++++++++++++++++')

    return results


def testfile_jsonl_totxt(in_path, out_path):
    w = open(out_path, 'w', encoding='utf-8')

    r_lines = jsonlines.open(in_path)
    for line in r_lines:
        w.write(line['summary'].strip() + '\n')


def testfile_csv_totxt(in_path, out_path):
    w = open(out_path, 'w', encoding='utf-8')
    f = open(in_path, 'r', encoding='utf-8')
    rows = csv.reader(f)
    for row in rows:
        w.write(row[0].strip() + '\n')


def FileRouge_google_verison(predictions_path, references_path):
    FR = FilesRouge()
    results = FR.get_scores(hyp_path=predictions_path, ref_path=references_path, avg=True)
    print(results)
    print(len(results))


def Rouge_line_test(predictions_path, references_path):
    f_1 = open(predictions_path, 'r', encoding='utf-8')
    lines_1 = f_1.readlines()

    f_2 = open(references_path, 'r', encoding='utf-8')
    lines_2 = f_2.readlines()

    scorer = rouge_scorer.RougeScorer(['rouge1', 'rougeL'], use_stemmer=False)

    fmeasure_all = 0

    for idx, lines in enumerate(zip(lines_1, lines_2)):
        scores = scorer.score(" ".join(lines[0]).strip(), " ".join(lines[1]).strip())
        print(scores)
        rouge1 = scores['rouge1']
        fmeasure = rouge1.fmeasure
        fmeasure_all += fmeasure

    print(fmeasure_all / (idx + 1))


def Rouge_line(references_path, predictions_path):
    predictions_list = []
    references_list = []
    pf = open(predictions_path, 'r', encoding='utf-8')
    r_lines = jsonlines.open(references_path)
    # print(r_lines)

    p_rows = csv.reader(pf)

    for row in p_rows:
        predictions_list.append(row[0])

    for line in r_lines:
        references_list.append(line['summary'].strip())

    scorer = rouge_scorer.RougeScorer(['rouge1', 'rouge2', 'rougeL'], use_stemmer=False)
    fmeasure_all_rouge1 = 0
    fmeasure_all_rouge2 = 0
    fmeasure_all_rougeL = 0

    for idx, lines in enumerate(zip(references_list, predictions_list)):
        # print(lines[0])
        # print(lines[1])
        scores = scorer.score(" ".join(lines[0]).strip(), " ".join(lines[1]).strip())
        # scores = scorer.score(lines[0].strip(), lines[1].strip())
        rouge1 = scores['rouge1']
        rouge1_fmeasure = rouge1.fmeasure
        fmeasure_all_rouge1 += rouge1_fmeasure

        rouge2 = scores['rouge2']
        rouge2_fmeasure = rouge2.fmeasure
        fmeasure_all_rouge2 += rouge2_fmeasure

        rougeL = scores['rougeL']
        rougeL_fmeasure = rougeL.fmeasure
        fmeasure_all_rougeL += rougeL_fmeasure




    rouge1 = fmeasure_all_rouge1 / (idx + 1)
    rouge2 = fmeasure_all_rouge2 / (idx + 1)
    rougeL = fmeasure_all_rougeL / (idx + 1)

    result = {'rouge1': rouge1, 'rouge2': rouge2, 'rougeL': rougeL}

    return result



if __name__ == '__main__':
    # enzh
    # enzh_results = Rouge(
    #     predictions_path=r'/data5/haoyun.xu/study/MT/Finetune_LLM_for_MT/infer/output/final/4_3090/summary/train_neuron/union/frzh/150000x2/4096_cut/38step/enzh_output.csv',
    #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/english-chinese_simplified/4096_cut_english-chinese_simplified_test.jsonl')
    #
    # # frzh
    # frzh_results = Rouge(
    #     predictions_path=r'/data5/haoyun.xu/study/MT/Finetune_LLM_for_MT/infer/output/final/4_3090/summary/train_neuron/union/frzh/150000x2/4096_cut/38step/frzh_output.csv',
    #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/french-chinese_simplified/4096_cut_french-chinese_simplified_test.jsonl')
    #
    # # # hizh
    # hizh_results = Rouge(
    #     predictions_path=r'/data5/haoyun.xu/study/MT/Finetune_LLM_for_MT/infer/output/final/4_3090/summary/train_neuron/union/frzh/150000x2/4096_cut/38step/hizh_output.csv',
    #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/hindi-chinese_simplified/4096_cut_hindi-chinese_simplified_test.jsonl')

    # enfr
    # enfr_results = Rouge(
    #     predictions_path=r'/data5/haoyun.xu/study/MT/Finetune_LLM_for_MT/infer/output/final/4_3090/summary/org/enfr/4096_cut/136step/enfr_output.csv',
    #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/english-french/4096_cut_english-french_test.jsonl')
    #
    # # hifr
    # hifr_results = Rouge(
    #     predictions_path=r'/data5/haoyun.xu/study/MT/Finetune_LLM_for_MT/infer/output/final/4_3090/summary/org/enfr/4096_cut/136step/hifr_output.csv',
    #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/hindi-french/4096_cut_hindi-french_test.jsonl')


    path_temp = '/data5/haoyun.xu/study/MT/Finetune_LLM_for_MT/infer/output/final/4_3090/summary/llama2_chat'

    enzh_results = Rouge_line(
        references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/english-chinese_simplified/4096_cut_english-chinese_simplified_test.jsonl',
        predictions_path=path_temp + r'/enzh_output_2.csv')

    print(enzh_results)

    print('---------------------------------')

    frzh_results = Rouge_line(
        references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/french-chinese_simplified/4096_cut_french-chinese_simplified_test.jsonl',
        predictions_path=path_temp + r'/frzh_output_.csv')

    print(frzh_results)

    print('---------------------------------')

    hizh_results = Rouge_line(
        references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/hindi-chinese_simplified/4096_cut_hindi-chinese_simplified_test.jsonl',
        predictions_path=path_temp + r'/hizh_output.csv')

    print(hizh_results)

    print('---------------------------------')


    # # enfr_results = Rouge_line(
    # #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/english-french/4096_cut_english-french_test.jsonl',
    # #     predictions_path=path_temp+r'/enfr_output.csv')
    # #
    # # print('---------------------------------')
    # #
    # # hifr_results = Rouge_line(
    # #     references_path=r'/data5/haoyun.xu/study/MI/MMMI/src/MI/experiment_setup/baselines/data/CrossSum/hindi-french/4096_cut_hindi-french_test.jsonl',
    # #     predictions_path=path_temp+r'/hifr_output.csv')
    #
    # wb = Workbook()
    # sheet = wb.active
    # sheet.cell(row=1, column=1).value = enzh_results['rouge1']
    # sheet.cell(row=1, column=2).value = enzh_results['rouge2']
    # sheet.cell(row=1, column=3).value = enzh_results['rougeL']
    #
    # sheet.cell(row=1, column=6).value = frzh_results['rouge1']
    # sheet.cell(row=1, column=7).value = frzh_results['rouge2']
    # sheet.cell(row=1, column=8).value = frzh_results['rougeL']
    #
    # sheet.cell(row=1, column=11).value = hizh_results['rouge1']
    # sheet.cell(row=1, column=12).value = hizh_results['rouge2']
    # sheet.cell(row=1, column=13).value = hizh_results['rougeL']

    # sheet.cell(row=1, column=1).value = enfr_results['rouge1']
    # sheet.cell(row=1, column=2).value = enfr_results['rouge2']
    # sheet.cell(row=1, column=3).value = enfr_results['rougeL']
    #
    # sheet.cell(row=1, column=6).value = hifr_results['rouge1']
    # sheet.cell(row=1, column=7).value = hifr_results['rouge2']
    # sheet.cell(row=1, column=8).value = hifr_results['rougeL']

    # wb.save('temp.xlsx')
